# main.py
import os
import platform
import xml.etree.ElementTree as et
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side

XLSX_DIRECTORY = "./xlsx"


def get_default_sequence_directory():
    """Get the default MA3 sequence directory based on the operating system."""
    if platform.system() == "Windows":
        return "C:\\ProgramData\\MALightingTechnology\\gma3_library\\datapools\\sequences\\"
    elif platform.system() == "Darwin":
        return f'{os.path.expanduser("~")}/MALightingTechnology/gma3_library/datapools/sequences/'
    return ""


def initialize_application():
    """Initialize the application and return menu dictionary."""
    default_dir = get_default_sequence_directory()
    if not os.path.isdir(default_dir):
        print(f"Error: Directory '{default_dir}' does not exist or is inaccessible.")
        exit(1)
    menu = {i + 1: file for i, file in enumerate(f for f in os.listdir(default_dir) if f.endswith(".xml"))}
    menu[len(menu) + 1] = "Exit"
    return menu, default_dir


def display_menu(menu):
    """Display a formatted menu to the console."""
    print("-" * 30)
    print("MENU\n" + "-" * 30)
    for key, value in menu.items():
        print(f"{key} -- {value}")


def process_cue(cue):
    """Process a single cue element and return its data."""
    number = float(cue.get("No", "0").strip()) if cue.get("No") else 0
    name = cue.get("Name", "")
    note = cue.get("Note", "")
    cuefadein = next((child.attrib['CueInFade'] for child in cue if "CueInFade" in child.attrib), "")
    cuefadeout = next((child.attrib['CueOutFade'] for child in cue if "CueOutFade" in child.attrib), "")
    trigtype = cue.get("TrigType", "Go+")
    if trigtype == "Time":
        trigtype += f" - {cue.get('TrigTime', '')}"
    elif trigtype == "Sound":
        trigtype += f" - {cue.get('TrigSound', '')}"
    return [number, name, cuefadein, cuefadeout, note, trigtype, ""]


def process_xml_file(file_path, filename):
    """Process XML file and return sequence data."""
    root = et.parse(file_path).getroot()
    version = root.attrib.get("DataVersion", "Unknown")

    sequence_element = root.find(".//Sequence")
    if sequence_element is not None and "Note" in sequence_element.attrib:
        sequence_note = sequence_element.attrib["Note"].replace("&#xD;", " ").strip()
    else:
        sequence_note = ""

    header = [["Num:", "Cue name:", "FadeIn:", "FadeOut", "Cue Note:", "Trig.Type/Param:", "Comment:"]]
    cues = [process_cue(cue) for cue in root.iter("Cue")]

    return header + cues, version, sequence_note, os.path.splitext(filename)[0]


def ensure_directory_exists(directory):
    """Ensure the specified directory exists, creating it if necessary."""
    if not os.path.exists(directory):
        print(f"Directory '{directory}' not found, creating it.")
        os.makedirs(directory)


def create_excel_file(tree_data, version, sequence_note, title):
    """Create and save an Excel file with sequence data."""
    ensure_directory_exists(XLSX_DIRECTORY)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # Applying styles
    ws["A1"].font, ws["A1"] = Font(size=24, italic=True), f"Sequence name: {title}"
    ws["A2"].font, ws["A2"] = Font(size=14, bold=True), f"MA3 program version: {version}"
    ws["A3"].font, ws["A3"] = Font(size=24, italic=True), "Sequence note:"
    ws["A4"].font, ws["A4"] = Font(size=14, italic=True), sequence_note
    ws.merge_cells("A1:D1")

    for row in tree_data:
        ws.append(row)

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                         bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=5, max_col=7):
        for cell in row:
            cell.border = thin_border
            cell.font = Font(size=16)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 20

    wb.save(f"{XLSX_DIRECTORY}/{title}.xlsx")
    wb.close()


def run_application():
    """Main application logic."""
    menu, default_dir = initialize_application()
    while True:
        display_menu(menu)
        try:
            selected_option = int(input("Choose an XML file to convert to XLSX: "))
            if selected_option not in menu:
                print("Invalid selection. Please choose a valid menu option.")
                continue
            if menu[selected_option] == "Exit":
                break
        except ValueError:
            print("Invalid input. Please enter a number.")
            continue
        file_path = os.path.join(default_dir, menu[selected_option])
        tree_data, version, sequence_note, title = process_xml_file(file_path, menu[selected_option])
        create_excel_file(tree_data, version, sequence_note, title)
        print("File conversion completed.\nRestarting...")


if __name__ == "__main__":
    run_application()
